"""
Main Flask Application
NCRP Complaint Organizer & Case View System
"""

from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
import os
from werkzeug.utils import secure_filename
from datetime import datetime
import json

from file_parser import FileParser
from data_processor import DataProcessor
from duplicate_detector import DuplicateDetector
from excel_generator import ExcelGenerator

app = Flask(__name__)
app.secret_key = 'ncrp_complaint_organizer_2024'  # Change in production
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['UPLOAD_FOLDER'] = 'data/uploads'

# Initialize modules
file_parser = FileParser()
data_processor = DataProcessor()
duplicate_detector = DuplicateDetector()
excel_generator = ExcelGenerator()

# Ensure directories exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs('data', exist_ok=True)

ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls', 'pdf'}


def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/')
def index():
    """Main page - file upload interface"""
    return render_template('index.html')


@app.route('/checklist')
def checklist():
    """Checklist page for complainants"""
    return render_template('checklist.html')


@app.route('/form')
def complaint_form():
    """Sample complaint form (demo only)"""
    return render_template('form.html')


@app.route('/admin')
def admin_dashboard():
    """Admin dashboard to view processed complaints"""
    # Load master register if exists
    master_file = 'data/master_register.xlsx'
    stats = {
        'total_complaints': 0,
        'file_exists': os.path.exists(master_file)
    }
    
    return render_template('admin.html', stats=stats)


@app.route('/upload', methods=['POST'])
def upload_file():
    """Handle file upload"""
    if 'file' not in request.files:
        flash('No file selected', 'error')
        return redirect(url_for('index'))
    
    file = request.files['file']
    
    if file.filename == '':
        flash('No file selected', 'error')
        return redirect(url_for('index'))
    
    if file and allowed_file(file.filename):
        try:
            # FIRST: Aggressively clean up any corrupted file
            cleanup_corrupted_file()
            
            # Save uploaded file with proper error handling
            filename = secure_filename(file.filename)
            if not filename:
                flash('Invalid filename. Please rename the file and try again.', 'error')
                return redirect(url_for('index'))
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{timestamp}_{filename}"
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            
            # Ensure upload directory exists
            os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
            
            # Save file and verify it was saved correctly
            file.save(file_path)
            
            # Verify file was saved (not 0 bytes)
            if not os.path.exists(file_path):
                flash('Error: File was not saved. Please try again.', 'error')
                return redirect(url_for('index'))
            
            file_size = os.path.getsize(file_path)
            if file_size == 0:
                # Clean up empty file
                try:
                    os.remove(file_path)
                except:
                    pass
                flash('Error: Uploaded file is empty (0 bytes). Please check the file and try again.', 'error')
                return redirect(url_for('index'))
            
            # Parse file with improved error handling
            raw_complaints = file_parser.parse_file(file_path)
            
            if not raw_complaints:
                flash('No complaints found in the file', 'error')
                return redirect(url_for('index'))
            
            # Process complaints
            processed_complaints = [data_processor.normalize_complaint(c) for c in raw_complaints]
            
            # Load existing master register
            master_complaints = load_master_register()
            
            # Add new complaints (avoid duplicates by complaint_id)
            existing_ids = {c.get('complaint_id') for c in master_complaints}
            new_complaints = [c for c in processed_complaints if c.get('complaint_id') not in existing_ids]
            
            if not new_complaints:
                flash('All complaints in the file already exist in the register', 'warning')
                return redirect(url_for('index'))
            
            # Add to master
            master_complaints.extend(new_complaints)
            
            # Categorize
            categorized = data_processor.categorize_complaints(master_complaints)
            
            # Find duplicates
            duplicate_groups = duplicate_detector.find_duplicates(master_complaints)
            formatted_duplicates = duplicate_detector.format_duplicate_groups(duplicate_groups)
            
            # Get high-value cases
            high_value_cases = data_processor.get_high_value_cases(master_complaints, threshold=50000)
            
            # Generate Excel
            excel_path = excel_generator.generate_excel(
                master_complaints,
                categorized,
                formatted_duplicates,
                high_value_cases
            )
            
            flash(f'Successfully processed {len(new_complaints)} new complaint(s). Total: {len(master_complaints)}', 'success')
            return redirect(url_for('download_excel'))
        
        except Exception as e:
            flash(f'Error processing file: {str(e)}', 'error')
            return redirect(url_for('index'))
    
    else:
        flash('Invalid file type. Please upload CSV, Excel, or PDF files only.', 'error')
        return redirect(url_for('index'))


def cleanup_corrupted_file():
    """Aggressively clean up corrupted Excel file"""
    master_file = 'data/master_register.xlsx'
    
    if not os.path.exists(master_file):
        return True  # No file to clean
    
    # Try multiple methods to validate and remove corrupted file
    try:
        # Method 1: Try zipfile validation
        import zipfile
        with zipfile.ZipFile(master_file, 'r') as zip_ref:
            if '[Content_Types].xml' not in zip_ref.namelist():
                raise KeyError("Missing [Content_Types].xml")
    except (KeyError, zipfile.BadZipFile, Exception):
        # File is corrupted, try to remove it
        print(f"Detected corrupted file: {master_file}")
        import time
        backup_name = f"data/master_register_corrupted_{int(time.time())}.xlsx"
        
        # Try rename first (safer)
        try:
            os.rename(master_file, backup_name)
            print(f"Successfully renamed corrupted file to: {backup_name}")
            return True
        except Exception:
            # Try delete
            try:
                os.remove(master_file)
                print(f"Successfully deleted corrupted file: {master_file}")
                return True
            except Exception as e:
                print(f"Could not remove corrupted file (may be locked): {str(e)}")
                # File is locked, but we'll work around it
                return False
    
    # File is valid
    try:
        from openpyxl import load_workbook
        test_wb = load_workbook(master_file, read_only=True)
        test_wb.close()
        return True
    except Exception:
        # File might be corrupted in a different way
        print(f"File failed openpyxl validation, attempting cleanup")
        import time
        backup_name = f"data/master_register_corrupted_{int(time.time())}.xlsx"
        try:
            os.rename(master_file, backup_name)
            return True
        except:
            try:
                os.remove(master_file)
                return True
            except:
                return False


@app.route('/submit_form', methods=['POST'])
def submit_form():
    """Handle complaint form submission"""
    try:
        # FIRST: Aggressively clean up any corrupted file
        cleanup_corrupted_file()
        
        # Get form data
        complaint_data = {
            'complaint_id': request.form.get('complaint_id', ''),
            'complaint_date': request.form.get('complaint_date', ''),
            'incident_date': request.form.get('incident_date', ''),
            'complainant_name': request.form.get('complainant_name', ''),
            'mobile': request.form.get('mobile', ''),
            'email': request.form.get('email', ''),
            'district': request.form.get('district', ''),
            'police_station': request.form.get('police_station', ''),
            'crime_type': request.form.get('crime_type', ''),
            'platform': request.form.get('platform', ''),
            'amount': request.form.get('amount', '0'),
            'status': request.form.get('status', 'Registered'),
            'description': request.form.get('description', '')
        }
        
        # Generate complaint ID if not provided
        if not complaint_data['complaint_id']:
            complaint_data['complaint_id'] = f"FORM_{datetime.now().strftime('%Y%m%d%H%M%S')}"
        
        # Process complaint
        processed_complaint = data_processor.normalize_complaint(complaint_data)
        
        # Load existing master register (should be safe now)
        # If file was corrupted, this will return empty list and we'll create new file
        master_complaints = load_master_register()
        
        # Ensure we have a list
        if not isinstance(master_complaints, list):
            master_complaints = []
        
        # Check for duplicate
        existing_ids = {c.get('complaint_id') for c in master_complaints}
        if processed_complaint.get('complaint_id') in existing_ids:
            flash('A complaint with this ID already exists', 'warning')
            return redirect(url_for('complaint_form'))
        
        # Add to master
        master_complaints.append(processed_complaint)
        
        # Categorize
        categorized = data_processor.categorize_complaints(master_complaints)
        
        # Find duplicates
        duplicate_groups = duplicate_detector.find_duplicates(master_complaints)
        formatted_duplicates = duplicate_detector.format_duplicate_groups(duplicate_groups)
        
        # Get high-value cases
        high_value_cases = data_processor.get_high_value_cases(master_complaints, threshold=50000)
        
        # Generate Excel
        excel_path = excel_generator.generate_excel(
            master_complaints,
            categorized,
            formatted_duplicates,
            high_value_cases
        )
        
        flash('Complaint submitted successfully! (This is a demo form only)', 'success')
        return redirect(url_for('download_excel'))
    
    except Exception as e:
        flash(f'Error submitting complaint: {str(e)}', 'error')
        return redirect(url_for('complaint_form'))


@app.route('/download')
def download_excel():
    """Download the generated Excel file"""
    excel_path = 'data/master_register.xlsx'
    
    # Check for backup files if main file doesn't exist or is corrupted
    if not os.path.exists(excel_path):
        # Look for backup files
        import glob
        backup_files = glob.glob('data/master_register_backup_*.xlsx')
        if backup_files:
            # Use the most recent backup
            backup_files.sort(reverse=True)
            excel_path = backup_files[0]
        else:
            flash('No register file found. Please upload a file first.', 'error')
            return redirect(url_for('index'))
    
    # Validate file is not corrupted before sending
    try:
        from openpyxl import load_workbook
        test_wb = load_workbook(excel_path, read_only=True)
        test_wb.close()
    except Exception:
        # File is corrupted, look for backup
        import glob
        backup_files = glob.glob('data/master_register_backup_*.xlsx')
        if backup_files:
            backup_files.sort(reverse=True)
            excel_path = backup_files[0]
            flash('Using backup file (original was corrupted)', 'warning')
        else:
            flash('Register file is corrupted. Please upload a new file.', 'error')
            return redirect(url_for('index'))
    
    return send_file(
        excel_path,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'master_register_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


@app.route('/api/stats')
def get_stats():
    """Get statistics about complaints"""
    master_complaints = load_master_register()
    
    stats = {
        'total_complaints': len(master_complaints),
        'by_crime_type': {},
        'by_status': {},
        'total_amount': sum(c.get('amount', 0) for c in master_complaints),
        'high_value_count': len([c for c in master_complaints if c.get('amount', 0) >= 50000])
    }
    
    # Count by crime type
    for complaint in master_complaints:
        crime_type = complaint.get('crime_type', 'Other')
        stats['by_crime_type'][crime_type] = stats['by_crime_type'].get(crime_type, 0) + 1
        
        status = complaint.get('status', 'Registered')
        stats['by_status'][status] = stats['by_status'].get(status, 0) + 1
    
    return jsonify(stats)


def load_master_register():
    """Load existing master register from Excel"""
    import pandas as pd
    from openpyxl import load_workbook
    from zipfile import BadZipFile
    import zipfile
    
    master_file = 'data/master_register.xlsx'
    
    if not os.path.exists(master_file):
        return []
    
    # First, validate the file is not corrupted using multiple methods
    # Catch all possible exceptions including KeyError for missing [Content_Types].xml
    file_is_valid = False
    try:
        # Method 1: Try to validate as a zip file first (Excel files are zip archives)
        with zipfile.ZipFile(master_file, 'r') as zip_ref:
            # Check if [Content_Types].xml exists
            if '[Content_Types].xml' not in zip_ref.namelist():
                raise KeyError("Missing [Content_Types].xml")
            file_is_valid = True
    except (KeyError, BadZipFile, zipfile.BadZipFile) as e:
        # File is corrupted (missing [Content_Types].xml or not a valid zip)
        error_msg = str(e)
        print(f"File is corrupted (zip validation failed: {error_msg}), attempting to remove: {master_file}")
        try:
            # Try to rename first (safer than delete)
            import time
            backup_name = f"data/master_register_corrupted_{int(time.time())}.xlsx"
            os.rename(master_file, backup_name)
            print(f"Renamed corrupted file to: {backup_name}")
        except Exception as rename_error:
            # If rename fails, try to delete
            try:
                os.remove(master_file)
                print(f"Deleted corrupted file: {master_file}")
            except Exception as del_error:
                print(f"Could not remove corrupted file: {str(del_error)}")
        return []
    except Exception as e:
        # Other error, try openpyxl validation
        try:
            test_wb = load_workbook(master_file, read_only=True)
            test_wb.close()
            file_is_valid = True
        except Exception as e2:
            # File is corrupted
            print(f"File is corrupted (openpyxl validation failed: {str(e2)}), attempting to remove: {master_file}")
            try:
                import time
                backup_name = f"data/master_register_corrupted_{int(time.time())}.xlsx"
                os.rename(master_file, backup_name)
                print(f"Renamed corrupted file to: {backup_name}")
            except Exception:
                try:
                    os.remove(master_file)
                    print(f"Deleted corrupted file: {master_file}")
                except:
                    pass
            return []
    
    # File is valid, try to read it with pandas
    if file_is_valid:
        try:
            # Try to read Master_Sheet, if it doesn't exist, return empty list
            df = pd.read_excel(master_file, sheet_name='Master_Sheet')
            # Convert NaN values to empty strings or None
            df = df.fillna('')
            return df.to_dict('records')
        except Exception as e:
            # If sheet doesn't exist or other error, return empty list
            print(f"Error reading master register with pandas: {str(e)}")
            return []
    
    return []


if __name__ == '__main__':
    print("=" * 60)
    print("NCRP Complaint Organizer & Case View System")
    print("=" * 60)
    print("\n⚠️  IMPORTANT: This is a DEMO system only")
    print("   - Does NOT access real NCRP portal")
    print("   - Works with uploaded files and sample data")
    print("   - All data stored locally")
    print("\n🚀 Starting server on http://localhost:5000")
    print("=" * 60)
    
    app.run(debug=True, host='127.0.0.1', port=5000)

