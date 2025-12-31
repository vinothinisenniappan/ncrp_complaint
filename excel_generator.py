"""
Excel Generator Module
Creates Excel files with multiple categorized sheets
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import zipfile


class ExcelGenerator:
    """Generate Excel files with categorized sheets"""
    
    def __init__(self, output_path='data/master_register.xlsx'):
        self.output_path = output_path
        self.ensure_data_directory()
    
    def ensure_data_directory(self):
        """Ensure data directory exists"""
        os.makedirs(os.path.dirname(self.output_path), exist_ok=True)
    
    def generate_excel(self, master_complaints, categorized_complaints, duplicate_groups, high_value_cases):
        """
        Generate Excel file with multiple sheets
        """
        # Ensure master_complaints is a list
        if not master_complaints:
            master_complaints = []
        
        # CRITICAL: Clean up corrupted file before attempting to write
        if os.path.exists(self.output_path):
            file_is_valid = False
            try:
                # Quick validation using zipfile (Excel files are zip archives)
                import zipfile
                with zipfile.ZipFile(self.output_path, 'r') as zip_ref:
                    if '[Content_Types].xml' in zip_ref.namelist():
                        file_is_valid = True
            except Exception:
                file_is_valid = False
            
            if not file_is_valid:
                # File is corrupted, try to remove it
                print(f"Detected corrupted file at: {self.output_path}")
                import time
                backup_name = f"data/master_register_corrupted_{int(time.time())}.xlsx"
                try:
                    os.rename(self.output_path, backup_name)
                    print(f"Renamed corrupted file to: {backup_name}")
                except Exception:
                    try:
                        os.remove(self.output_path)
                        print(f"Deleted corrupted file: {self.output_path}")
                    except Exception as e:
                        # If we can't remove it, create with timestamp to avoid conflict
                        print(f"Could not remove corrupted file (locked?): {str(e)}")
                        import time
                        self.output_path = f"data/master_register_{int(time.time())}.xlsx"
                        print(f"Will create new file at: {self.output_path}")
        
        # Always create master sheet first (even if empty)
        # Use mode='w' to overwrite any existing file
        # If file is locked or corrupted, create a new one with timestamp
        original_path = self.output_path
        
        def write_all_sheets(writer):
            """Helper function to write all sheets"""
            # Sheet 1: Master Sheet (All Complaints) - ALWAYS create this
            if master_complaints:
                self._write_master_sheet(writer, master_complaints)
            else:
                # Create empty master sheet with headers
                self._write_empty_master_sheet(writer)
            
            # Sheet 2: Crime-Type-wise sheets
            for category, complaints in categorized_complaints.items():
                if complaints:  # Only create sheet if there are complaints
                    self._write_category_sheet(writer, category, complaints)
            
            # Sheet: High Value Cases
            if high_value_cases:
                self._write_high_value_sheet(writer, high_value_cases)
            
            # Sheet: Possible Duplicates
            if duplicate_groups:
                self._write_duplicate_sheet(writer, duplicate_groups)
        
        try:
            with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
                write_all_sheets(writer)
        except (PermissionError, IOError, OSError, KeyError) as e:
            # File might be locked or corrupted, try with a timestamped name
            import time
            timestamp = int(time.time())
            backup_path = f"data/master_register_backup_{timestamp}.xlsx"
            self.output_path = backup_path
            print(f"Original file locked/corrupted, creating backup: {backup_path}")
            # Retry with new path
            try:
                with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
                    write_all_sheets(writer)
            except Exception as e2:
                print(f"Error creating backup file: {str(e2)}")
                raise
        except Exception as e:
            # If writing fails, try to create a minimal valid file
            import traceback
            print(f"Error writing Excel: {str(e)}")
            print(traceback.format_exc())
            # Create a simple valid Excel file
            self._create_minimal_excel(master_complaints)
        
        # Apply formatting ONLY if file is a valid Excel archive
        try:
            if self.is_valid_excel(self.output_path):
                try:
                    self._apply_formatting(self.output_path)
                except Exception as e:
                    # Formatting is optional; log and continue
                    print(f"Formatting skipped: {e}")
            else:
                print("Excel file is invalid. Skipping formatting.")
        except Exception as e:
            # Validation failed unexpectedly; skip formatting
            print(f"Excel validation skipped due to error: {e}")

        return self.output_path
    
    
    def is_valid_excel(self, path):
        """Verify the Excel file is a valid XLSX (ZIP with content types)."""
        try:
            with zipfile.ZipFile(path, 'r') as z:
                names = z.namelist()
                return ('[Content_Types].xml' in names) and any(name.startswith('xl/') for name in names)
        except Exception:
            return False

    def _create_minimal_excel(self, master_complaints):
        """Create a minimal valid Excel file if main creation fails"""
        try:
            if master_complaints:
                df = pd.DataFrame(master_complaints)
            else:
                df = pd.DataFrame(columns=['complaint_id', 'complaint_date', 'complainant_name'])
            df.to_excel(self.output_path, sheet_name='Master_Sheet', index=False)
        except Exception as e:
            print(f"Error creating minimal Excel: {str(e)}")
            raise
    
    def _write_master_sheet(self, writer, complaints):
        """Write master sheet with all complaints"""
        if not complaints:
            self._write_empty_master_sheet(writer)
            return
        
        df = pd.DataFrame(complaints)
        
        # Reorder columns for better readability
        column_order = [
            'complaint_id', 'complaint_date', 'incident_date', 'complainant_name',
            'mobile', 'email', 'district', 'police_station', 'crime_type',
            'platform', 'amount', 'status', 'description'
        ]
        
        # Only include columns that exist
        available_columns = [col for col in column_order if col in df.columns]
        if available_columns:
            df = df[available_columns]
            df.to_excel(writer, sheet_name='Master_Sheet', index=False)
        else:
            self._write_empty_master_sheet(writer)
    
    def _write_empty_master_sheet(self, writer):
        """Write empty master sheet with headers only"""
        headers = [
            'complaint_id', 'complaint_date', 'incident_date', 'complainant_name',
            'mobile', 'email', 'district', 'police_station', 'crime_type',
            'platform', 'amount', 'status', 'description'
        ]
        df = pd.DataFrame(columns=headers)
        df.to_excel(writer, sheet_name='Master_Sheet', index=False)
    
    def _write_category_sheet(self, writer, category, complaints):
        """Write category-specific sheet"""
        df = pd.DataFrame(complaints)
        
        # Reorder columns
        column_order = [
            'complaint_id', 'complaint_date', 'complainant_name',
            'mobile', 'email', 'district', 'amount', 'status'
        ]
        
        available_columns = [col for col in column_order if col in df.columns]
        df = df[available_columns]
        
        # Clean sheet name (Excel has 31 char limit)
        sheet_name = category[:31] if len(category) > 31 else category
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    def _write_high_value_sheet(self, writer, complaints):
        """Write high-value cases sheet"""
        df = pd.DataFrame(complaints)
        
        column_order = [
            'complaint_id', 'complaint_date', 'complainant_name',
            'mobile', 'email', 'district', 'crime_type', 'platform',
            'amount', 'status'
        ]
        
        available_columns = [col for col in column_order if col in df.columns]
        df = df[available_columns]
        
        # Sort by amount descending
        if 'amount' in df.columns:
            df = df.sort_values('amount', ascending=False)
        
        df.to_excel(writer, sheet_name='High_Value_Cases', index=False)
    
    def _write_duplicate_sheet(self, writer, duplicate_groups):
        """Write possible duplicates sheet"""
        # Flatten duplicate groups into list
        all_duplicates = []
        for group in duplicate_groups:
            for complaint in group:
                all_duplicates.append(complaint)
        
        if not all_duplicates:
            return
        
        df = pd.DataFrame(all_duplicates)
        
        column_order = [
            'duplicate_group_id', 'complaint_id', 'complaint_date',
            'complainant_name', 'mobile', 'email', 'amount',
            'crime_type', 'platform', 'match_reason', 'group_size'
        ]
        
        available_columns = [col for col in column_order if col in df.columns]
        df = df[available_columns]
        
        df.to_excel(writer, sheet_name='Possible_Duplicates', index=False)
    
    def _apply_formatting(self, file_path):
        """Apply formatting to Excel file"""
        try:
            # Validate file exists and is readable
            if not os.path.exists(file_path):
                print(f"File does not exist for formatting: {file_path}")
                return
            
            wb = load_workbook(file_path)
        except (KeyError, OSError, IOError) as e:
            # File might be corrupted or locked - skip formatting but don't fail
            print(f"Warning: Could not load workbook for formatting (file may be locked): {str(e)}")
            return
        except Exception as e:
            print(f"Error loading workbook for formatting: {str(e)}")
            # Don't raise - formatting is optional
            return
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Skip if sheet is empty (no rows)
            if ws.max_row == 0:
                continue
            
            # Format header row (row 1)
            if ws.max_row >= 1:
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50) if max_length > 0 else 15
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Apply borders to all cells
            if ws.max_row >= 1:
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.border = border
                        if cell.row > 1:  # Not header
                            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Freeze header row (only if there's data)
            if ws.max_row > 1:
                ws.freeze_panes = 'A2'
        
        wb.save(file_path)

